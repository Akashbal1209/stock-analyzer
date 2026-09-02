from flask import Flask, render_template, jsonify, request
from flask_cors import CORS
import threading
import json
from datetime import datetime
import traceback
import numpy as np
import pandas as pd
import time

# Import configuration
try:
    from config import EXCEL_FILE_PATH, SERVER_HOST, SERVER_PORT, DEBUG_MODE
except ImportError:
    EXCEL_FILE_PATH = "Indian_Stocks_Complete_Market_Cap_Classification.xlsx"
    SERVER_HOST = "0.0.0.0"
    SERVER_PORT = 8080
    DEBUG_MODE = False

# Import existing backend modules
from stock_loader import StockLoader
from data_fetcher import DataFetcher
from signal_engine import SignalEngine
from comprehensive_analyzer import ComprehensiveAnalyzer
from backtester import AdvancedBacktester
from backtesting import SimpleBacktester

app = Flask(__name__)
CORS(app)

# Initialize backend modules
print("\n" + "=" * 60)
print("   STOCK SIGNAL ANALYZER API - Initializing...")
print("=" * 60 + "\n")

stock_loader = StockLoader(EXCEL_FILE_PATH)
data_fetcher = DataFetcher()
signal_engine = SignalEngine()
comprehensive_analyzer = ComprehensiveAnalyzer()
backtester = AdvancedBacktester()
simple_backtester = SimpleBacktester()

# ==================== SIGNAL CACHE ====================
signal_cache = {}
cache_loading = False
cache_loaded = False

def preload_db_stocks_signals():
    """Preload signals for DB stocks on startup - uses comprehensive_analyzer"""
    global signal_cache, cache_loading, cache_loaded
    
    cache_loading = True
    print("\n" + "=" * 60)
    print("   PRELOADING DB STOCKS SIGNALS...")
    print("=" * 60)
    
    total = len(stock_loader.db_stocks)
    success = 0
    failed = 0
    
    for i, stock in enumerate(stock_loader.db_stocks):
        symbol = stock['symbol']
        try:
            print(f"   [{i+1}/{total}] Loading {symbol}...", end=" ")
            
            # Use 2y data - same as detail page
            df = data_fetcher.fetch_data(symbol, period='2y')
            
            if df is not None and len(df) >= 50:
                # Use comprehensive_analyzer - SAME as stock detail page
                result = comprehensive_analyzer.analyze(symbol, df)
                
                if result and 'signal' in result:
                    signal_cache[symbol] = {
                        'signal': result['signal'],
                        'overall_score': result.get('overall_score', 0),
                        'current_price': result.get('current_price'),
                        'updated_at': datetime.now().strftime('%H:%M')
                    }
                    print(f"{result['signal']} (Score: {result.get('overall_score', 0)})")
                    success += 1
                else:
                    signal_cache[symbol] = {'signal': 'ERROR', 'error': 'Analysis failed'}
                    print("ERROR")
                    failed += 1
            else:
                signal_cache[symbol] = {'signal': 'NO_DATA'}
                print("NO DATA")
                failed += 1
                
        except Exception as e:
            signal_cache[symbol] = {'signal': 'ERROR', 'error': str(e)}
            print(f"FAILED: {e}")
            failed += 1
    
    cache_loading = False
    cache_loaded = True
    
    print("\n" + "=" * 60)
    print(f"   PRELOAD COMPLETE: {success} success, {failed} failed")
    print("=" * 60 + "\n")

# Start preloading in background thread
def start_preload():
    time.sleep(2)  # Wait for server to start
    preload_db_stocks_signals()

threading.Thread(target=start_preload, daemon=True).start()

print("\n[OK] All modules loaded successfully!")
print("[INFO] DB Stocks signals will preload in background...")
print("=" * 60)


# ==================== ROUTES ====================

@app.route('/')
def index():
    """Serve the main GUI"""
    return render_template('index.html')


@app.route('/api/stats')
def get_stats():
    """Get database statistics"""
    try:
        return jsonify({
            'success': True,
            'data': {
                'total_stocks': len(stock_loader.all_stocks),
                'db_stocks': len(stock_loader.db_stocks),
                'large_cap': len(stock_loader.large_cap),
                'mid_cap': len(stock_loader.mid_cap),
                'small_cap': len(stock_loader.small_cap)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/category/<category>')
def get_category_stocks(category):
    """Get all stocks in a category"""
    try:
        stocks = []
        if category == 'large_cap':
            stocks = stock_loader.large_cap
        elif category == 'mid_cap':
            stocks = stock_loader.mid_cap
        elif category == 'small_cap':
            stocks = stock_loader.small_cap
        elif category == 'db_stocks':
            stocks = stock_loader.db_stocks
        else:
            return jsonify({'success': False, 'error': 'Invalid category'})
        
        return jsonify({
            'success': True,
            'data': stocks,
            'count': len(stocks)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== DB STOCKS MANAGEMENT ====================

@app.route('/api/db-stocks')
def get_db_stocks():
    """Get all stocks in user's DB_Stocks list with signals"""
    try:
        stocks_with_signals = []
        
        for stock in stock_loader.db_stocks:
            symbol = stock['symbol']
            stock_data = stock.copy()
            
            # Add cached signal if available
            if symbol in signal_cache:
                cached = signal_cache[symbol]
                stock_data['signal'] = cached.get('signal', 'N/A')
                stock_data['confidence'] = cached.get('confidence', 0)
                stock_data['entry_price'] = cached.get('entry_price')
                stock_data['target'] = cached.get('target')
                stock_data['stop_loss'] = cached.get('stop_loss')
            else:
                stock_data['signal'] = 'LOADING' if cache_loading else 'N/A'
                stock_data['confidence'] = 0
            
            stocks_with_signals.append(stock_data)
        
        # Sort: BUY first, then SELL, then HOLD, then others
        signal_order = {'BUY': 0, 'SELL': 1, 'HOLD': 2, 'LOADING': 3, 'N/A': 4, 'ERROR': 5, 'NO_DATA': 6}
        stocks_with_signals.sort(key=lambda x: (signal_order.get(x.get('signal', 'N/A'), 99), -x.get('confidence', 0)))
        
        return jsonify({
            'success': True,
            'data': stocks_with_signals,
            'count': len(stocks_with_signals),
            'cache_status': 'loaded' if cache_loaded else 'loading' if cache_loading else 'pending'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/db-stocks/refresh')
def refresh_db_stocks_signals():
    """Refresh signals for all DB stocks"""
    try:
        threading.Thread(target=preload_db_stocks_signals, daemon=True).start()
        return jsonify({
            'success': True,
            'message': 'Refresh started in background'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/db-stocks/add', methods=['POST'])
def add_db_stock():
    """Add a stock to user's DB_Stocks list"""
    try:
        data = request.get_json()
        symbol = data.get('symbol', '').strip().upper()
        
        if not symbol:
            return jsonify({'success': False, 'error': 'Symbol is required'})
        
        existing = [s for s in stock_loader.db_stocks if s['symbol'] == symbol]
        if existing:
            return jsonify({'success': False, 'error': f'{symbol} already exists'})
        
        new_stock = {
            'symbol': symbol,
            'name': symbol,
            'category': 'DB_Stocks',
            'sector': 'Your Preferred',
            'market_cap': 'N/A'
        }
        stock_loader.db_stocks.append(new_stock)
        stock_loader.all_stocks.insert(0, new_stock)
        
        save_db_stocks_to_excel()
        
        return jsonify({
            'success': True,
            'message': f'{symbol} added successfully',
            'data': new_stock
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/db-stocks/remove', methods=['POST'])
def remove_db_stock():
    """Remove a stock from user's DB_Stocks list"""
    try:
        data = request.get_json()
        symbol = data.get('symbol', '').strip().upper()
        
        if not symbol:
            return jsonify({'success': False, 'error': 'Symbol is required'})
        
        stock_loader.db_stocks = [s for s in stock_loader.db_stocks if s['symbol'] != symbol]
        stock_loader.all_stocks = [s for s in stock_loader.all_stocks 
                                    if not (s['symbol'] == symbol and s['category'] == 'DB_Stocks')]
        
        save_db_stocks_to_excel()
        
        return jsonify({
            'success': True,
            'message': f'{symbol} removed successfully'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


def save_db_stocks_to_excel():
    """Save DB_Stocks list back to Excel file"""
    try:
        import openpyxl
        from openpyxl import load_workbook
        
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb['DB_Stocks']
        
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).value = None
            ws.cell(row=row, column=2).value = None
        
        for idx, stock in enumerate(stock_loader.db_stocks, start=2):
            ws.cell(row=idx, column=1).value = idx - 1
            ws.cell(row=idx, column=2).value = stock['symbol']
        
        wb.save(EXCEL_FILE_PATH)
        print(f"[OK] Saved {len(stock_loader.db_stocks)} stocks to Excel")
        
    except Exception as e:
        print(f"[WARN] Could not save to Excel: {e}")


# ==================== SEARCH ====================

@app.route('/api/search')
def search_stocks():
    """Search for stocks"""
    try:
        query = request.args.get('q', '').strip()
        if not query:
            return jsonify({'success': False, 'error': 'Please enter a search term'})
        
        results = stock_loader.search_stock(query)
        
        if not results:
            return jsonify({
                'success': True,
                'data': [],
                'message': f"'{query}' not found. You can still analyze using '{query.upper()}' as symbol.",
                'manual_symbol': query.upper()
            })
        
        return jsonify({
            'success': True,
            'data': results
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/analyze/<symbol>')
def analyze_stock(symbol):
    """Run COMPREHENSIVE analysis on a stock"""
    try:
        symbol = symbol.strip().upper()
        
        df = data_fetcher.fetch_data(symbol, period='2y')
        
        if df is None or len(df) < 50:
            return jsonify({
                'success': False,
                'error': f'Could not fetch sufficient data for {symbol}'
            })
        
        result = comprehensive_analyzer.analyze(symbol, df)
        
        if not result:
            return jsonify({
                'success': False,
                'error': f'Error analyzing {symbol}'
            })
        
        stock_info = stock_loader.get_stock_info(symbol)
        result = convert_to_serializable(result)
        
        return jsonify({
            'success': True,
            'data': {
                'symbol': symbol,
                'stock_info': stock_info,
                'analysis': result
            }
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/quick-signal/<symbol>')
def quick_signal(symbol):
    """Get quick signal using signal_engine"""
    try:
        symbol = symbol.strip().upper()
        
        df = data_fetcher.fetch_data(symbol, period='1y')
        
        if df is None or len(df) < 50:
            return jsonify({
                'success': False,
                'error': f'Could not fetch data for {symbol}'
            })
        
        result = signal_engine.analyze(df)
        
        if 'error' in result:
            return jsonify({
                'success': False,
                'error': result['error']
            })
        
        result = convert_to_serializable(result)
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/backtest/<symbol>')
def backtest_stock(symbol):
    """Run backtest (advanced version)"""
    try:
        symbol = symbol.strip().upper()
        date_str = request.args.get('date', '')
        
        if not date_str:
            return jsonify({
                'success': False,
                'error': 'Please provide a date (YYYY-MM-DD)'
            })
        
        df = data_fetcher.fetch_data(symbol, period='2y')
        
        if df is None or len(df) < 50:
            return jsonify({
                'success': False,
                'error': f'Could not fetch data for {symbol}'
            })
        
        result = backtester.backtest_single_date(df, date_str, symbol)
        
        if 'error' in result:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Backtest failed')
            })
        
        result = convert_to_serializable(result)
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== SIMPLE BACKTEST ====================

@app.route('/api/simple-backtest/<symbol>')
def simple_backtest_stock(symbol):
    """
    Simple Backtest - Easy to understand
    
    Shows:
    - What signal was generated on that date (BUY/SELL/HOLD)
    - Entry price, Target, and Stop Loss
    - Result - Was the signal successful or not
    
    Usage: /api/simple-backtest/RELIANCE?date=2024-06-15
    """
    try:
        symbol = symbol.strip().upper()
        date_str = request.args.get('date', '')
        
        if not date_str:
            return jsonify({
                'success': False,
                'error': 'Please provide date in format: YYYY-MM-DD (e.g., 2024-06-15)'
            })
        
        # Fetch 2 years data for proper analysis
        df = data_fetcher.fetch_data(symbol, period='2y')
        
        if df is None or len(df) < 50:
            return jsonify({
                'success': False,
                'error': f'Could not fetch data for {symbol}'
            })
        
        # Run simple backtest
        result = simple_backtester.run_backtest(df, date_str, symbol)
        
        if 'error' in result:
            return jsonify({
                'success': False,
                'error': result['error']
            })
        
        result = convert_to_serializable(result)
        
        return jsonify({
            'success': True,
            'data': result
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==================== GET AVAILABLE DATES ====================

@app.route('/api/available-dates/<symbol>')
def get_available_dates(symbol):
    """
    Get available trading dates for backtesting
    
    Returns list of dates where backtest can be run
    """
    try:
        symbol = symbol.strip().upper()
        
        df = data_fetcher.fetch_data(symbol, period='2y')
        
        if df is None or len(df) < 50:
            return jsonify({
                'success': False,
                'error': f'Could not fetch data for {symbol}'
            })
        
        # Get dates where backtest is possible (need 50 days before)
        df['Date'] = pd.to_datetime(df['Date'])
        available_dates = df['Date'].iloc[50:].tolist()
        
        # Convert to strings
        date_strings = [d.strftime('%Y-%m-%d') for d in available_dates]
        
        return jsonify({
            'success': True,
            'data': {
                'symbol': symbol,
                'total_dates': len(date_strings),
                'earliest_date': date_strings[0] if date_strings else None,
                'latest_date': date_strings[-1] if date_strings else None,
                'dates': date_strings[-90:]  # Last 90 dates for quick access
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/current-price/<symbol>')
def get_current_price(symbol):
    """Get current price"""
    try:
        symbol = symbol.strip().upper()
        price = data_fetcher.get_current_price(symbol)
        
        if price is None:
            return jsonify({
                'success': False,
                'error': f'Could not get price for {symbol}'
            })
        
        return jsonify({
            'success': True,
            'data': {'symbol': symbol, 'price': round(float(price), 2)}
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/stock-info/<symbol>')
def get_stock_info(symbol):
    """Get stock info"""
    try:
        symbol = symbol.strip().upper()
        info = data_fetcher.get_stock_info(symbol)
        
        if info is None:
            return jsonify({
                'success': False,
                'error': f'Could not get info for {symbol}'
            })
        
        return jsonify({
            'success': True,
            'data': info
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== HELPER FUNCTIONS ====================

def convert_to_serializable(obj):
    """Convert non-JSON-serializable objects to JSON-compatible types"""
    if obj is None:
        return None
    
    if isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    
    elif isinstance(obj, list):
        return [convert_to_serializable(item) for item in obj]
    
    elif isinstance(obj, tuple):
        return [convert_to_serializable(item) for item in obj]
    
    elif isinstance(obj, (datetime,)):
        return obj.strftime('%Y-%m-%d')
    
    elif hasattr(obj, 'strftime'):
        return obj.strftime('%Y-%m-%d')
    
    elif isinstance(obj, np.ndarray):
        return convert_to_serializable(obj.tolist())
    
    elif isinstance(obj, (np.integer, np.int64, np.int32, np.int16, np.int8)):
        return int(obj)
    
    elif isinstance(obj, (np.floating, np.float64, np.float32, np.float16)):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return float(obj)
    
    elif isinstance(obj, np.bool_):
        return bool(obj)
    
    elif isinstance(obj, pd.Series):
        return convert_to_serializable(obj.tolist())
    
    elif isinstance(obj, pd.DataFrame):
        return convert_to_serializable(obj.to_dict('records'))
    
    elif isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    
    elif hasattr(obj, 'item'):
        try:
            val = obj.item()
            return convert_to_serializable(val)
        except (ValueError, AttributeError):
            if hasattr(obj, 'tolist'):
                return convert_to_serializable(obj.tolist())
            return str(obj)
    
    else:
        return obj


# ==================== RUN SERVER ====================

if __name__ == '__main__':
    import webbrowser
    import time
    
    def open_browser():
        """Open browser after server starts"""
        time.sleep(1.5)
        webbrowser.open(f'http://127.0.0.1:{SERVER_PORT}')
    
    print("\n" + "=" * 60)
    print("   STOCK ANALYZER PRO - Web GUI")
    print("=" * 60)
    print(f"\n   Starting web server...")
    print(f"   URL: http://127.0.0.1:{SERVER_PORT}")
    print("\n   Simple Backtest API:")
    print(f"   /api/simple-backtest/SYMBOL?date=YYYY-MM-DD")
    print("\n   Press Ctrl+C to stop the server")
    print("=" * 60 + "\n")
    
    # Auto-open browser
    threading.Thread(target=open_browser, daemon=True).start()
    
    try:
        app.run(host=SERVER_HOST, port=SERVER_PORT, debug=DEBUG_MODE, threaded=True)
    except KeyboardInterrupt:
        print("\n\n   Server stopped. Goodbye!")
